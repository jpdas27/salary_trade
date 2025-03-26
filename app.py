import yfinance as yf
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta
import time
import requests
from requests.adapters import HTTPAdapter
from requests.packages.urllib3.util.retry import Retry

pd.set_option('display.float_format', '{:.2f}'.format)

def is_last_thursday():
    today = datetime.today()
    # Find last day of the month
    last_day = today.replace(day=1) + timedelta(days=32)
    last_day = last_day.replace(day=1) - timedelta(days=1)

    # Find the last Thursday
    while last_day.weekday() != 3:  # 3 = Thursday
        last_day -= timedelta(days=1)

    return today.date() == last_day.date()

def create_session():
    session = requests.Session()
    retries = Retry(
        total=5,
        backoff_factor=1,
        status_forcelist=[429, 500, 502, 503, 504]
    )
    adapter = HTTPAdapter(max_retries=retries)
    session.mount("http://", adapter)
    session.mount("https://", adapter)
    return session

stocks_list = [
    "SRF.NS", "BAJFINANCE.NS", "NAVINFLUOR.NS", "MARUTI.NS", "UPL.NS", "IDEA.NS", "SBICARD.NS", "MANAPPURAM.NS",
    "TATACONSUM.NS", "INDUSINDBK.NS", "KOTAKBANK.NS", "BAJAJFINSV.NS", "MUTHOOTFIN.NS", "BHARTIARTL.NS", "LAURUSLABS.NS",
    "CHOLAFIN.NS", "JKCEMENT.NS", "EICHERMOT.NS", "LTTS.NS", "SHREECEM.NS", "WIPRO.NS", "M&MFIN.NS", "MGL.NS", "BERGEPAINT.NS",
    "BRITANNIA.NS", "AARTIIND.NS", "M&M.NS", "DABUR.NS", "UBL.NS", "ABBOTINDIA.NS", "DMART.NS", "BIOCON.NS", "CHAMBLFERT.NS",
    "COROMANDEL.NS", "TVSMOTOR.NS", "INDUSTOWER.NS", "MARICO.NS", "AUBANK.NS", "NYKAA.NS", "SBILIFE.NS", "NESTLEIND.NS",
    "DIVISLAB.NS", "JSWSTEEL.NS", "IGL.NS", "LTF.NS", "HDFCLIFE.NS", "RELIANCE.NS", "BAJAJ-AUTO.NS", "RBLBANK.NS", "CIPLA.NS",
    "HINDUNILVR.NS", "INFY.NS", "TITAN.NS", "BSE.NS", "ULTRACEMCO.NS", "ZYDUSLIFE.NS", "LUPIN.NS", "JUBLFOOD.NS", "INDIGO.NS",
    "IDFCFIRSTB.NS", "DALBHARAT.NS", "ONGC.NS", "GRASIM.NS", "BATAINDIA.NS", "TECHM.NS", "GODREJCP.NS", "ICICIBANK.NS",
    "ASIANPAINT.NS", "GRANULES.NS", "DRREDDY.NS", "GLENMARK.NS", "ADANIENT.NS", "ICICIGI.NS", "IEX.NS", "SUNPHARMA.NS",
    "BALKRISIND.NS", "KPITTECH.NS", "HDFCBANK.NS", "MFSL.NS", "ESCORTS.NS", "PIDILITIND.NS", "GNFC.NS", "HINDALCO.NS",
    "BANKINDIA.NS", "ADANIPORTS.NS", "IRCTC.NS", "CUB.NS", "TCS.NS", "TORNTPHARM.NS", "INDIAMART.NS", "AUROPHARMA.NS",
    "HEROMOTOCO.NS", "PERSISTENT.NS", "LTIM.NS", "OIL.NS", "COLPAL.NS", "IPCALAB.NS", "ADANIENSOL.NS", "PHOENIXLTD.NS",
    "INDIANB.NS", "HAVELLS.NS", "ASHOKLEY.NS", "PNB.NS", "LALPATHLAB.NS", "TATASTEEL.NS", "UNITDSPR.NS", "HCLTECH.NS",
    "MAXHEALTH.NS", "PETRONET.NS", "AXISBANK.NS", "UNIONBANK.NS", "AMBUJACEM.NS", "MPHASIS.NS", "COALINDIA.NS", "FEDERALBNK.NS",
    "NHPC.NS", "ITC.NS", "ACC.NS", "LICHSGFIN.NS", "PAGEIND.NS", "POONAWALLA.NS", "IRB.NS", "SHRIRAMFIN.NS", "YESBANK.NS",
    "TATAMOTORS.NS", "INDHOTEL.NS", "COFORGE.NS", "JINDALSTEL.NS", "APLAPOLLO.NS", "GMRAIRPORT.NS", "NBCC.NS", "NAUKRI.NS",
    "VEDL.NS", "SBIN.NS", "APOLLOHOSP.NS", "NTPC.NS", "RAMCOCEM.NS", "BEL.NS", "PIIND.NS", "CANBK.NS", "IOC.NS", "LT.NS",
    "BPCL.NS", "TATAELXSI.NS", "BANDHANBNK.NS", "ABFRL.NS", "LICI.NS", "ICICIPRULI.NS", "CROMPTON.NS", "ALKEM.NS", "TATACOMM.NS",
    "BSOFT.NS", "SOLARINDS.NS", "CONCOR.NS", "METROPOLIS.NS", "HDFCAMC.NS", "SUNTV.NS", "MCX.NS", "TATAPOWER.NS", "DEEPAKNTR.NS",
    "GUJGASLTD.NS", "MRF.NS", "BANKBARODA.NS", "ABCAPITAL.NS", "LODHA.NS", "BHARATFORG.NS", "DLF.NS", "JSL.NS", "ATGL.NS",
    "SAIL.NS", "NMDC.NS", "IRFC.NS", "DIXON.NS", "SYNGENE.NS", "VBL.NS", "ADANIGREEN.NS", "HAL.NS", "HINDPETRO.NS", "HINDCOPPER.NS",
    "CUMMINSIND.NS", "EXIDEIND.NS", "BHEL.NS", "TATACHEM.NS", "KEI.NS", "PEL.NS", "POWERGRID.NS", "SUPREMEIND.NS", "SONACOMS.NS",
    "GAIL.NS", "TORNTPOWER.NS", "ATUL.NS", "NATIONALUM.NS", "HUDCO.NS", "SJVN.NS", "ASTRAL.NS", "MOTHERSON.NS", "CANFINHOME.NS",
    "DELHIVERY.NS", "APOLLOTYRE.NS", "OBEROIRLTY.NS", "BOSCHLTD.NS", "PFC.NS", "PAYTM.NS", "CGPOWER.NS", "RECLTD.NS", "HFCL.NS",
    "ZOMATO.NS", "POLYCAB.NS", "TRENT.NS", "TIINDIA.NS", "VOLTAS.NS", "POLICYBZR.NS", "PRESTIGE.NS", "CYIENT.NS", "ABB.NS", "OFSS.NS",
    "ANGELONE.NS", "JIOFIN.NS", "PVRINOX.NS", "JSWENERGY.NS", "SIEMENS.NS", "GODREJPROP.NS", "KALYANKJIL.NS", "CESC.NS", "CDSL.NS",
    "CAMS.NS"
]

def find_60days_historical_avg(symbol, session):
    try:
        stock_data = yf.download(symbol, period='90d', session=session)
        
        if len(stock_data) < 60:
            print(f"{symbol}: Insufficient data for 60-day calculation")
            return None
            
        filter_stock_data = stock_data.tail(60).copy()
        filter_stock_data['average_price'] = (
            filter_stock_data['Open'] + 
            filter_stock_data['High'] + 
            filter_stock_data['Low'] + 
            filter_stock_data['Close']
        ) / 4
        
        return round(filter_stock_data['average_price'].mean(), 2)
        
    except Exception as e:
        print(f"Error in 60-day avg for {symbol}: {e}")
        return None

def fetch_current_price(stock_obj):
    try:
        return round(stock_obj.fast_info["last_price"], 2)
    except Exception as e:
        print(f"Error fetching current price: {e}")
        return None

def fetch_x_SMA_value(stock_obj, days_to_cal):
    try:
        df = stock_obj.history(period="3mo")
        sma_days = f'SMA_{days_to_cal}'
        df[sma_days] = df["Close"].rolling(window=days_to_cal).mean()
        return round(df[sma_days].dropna().iloc[-1], 2)
    except Exception as e:
        print(f"Error calculating SMA {days_to_cal}: {e}")
        return None

def cal_diff(cur_price, avg_price):
    if avg_price is None or cur_price is None:
        return None
    try:
        return round(((cur_price - avg_price) / avg_price) * 100, 2)
    except ZeroDivisionError:
        return None

def main():
    # Get F&O stocks
    fno_symbols = stocks_list
    if not fno_symbols:
        print("Failed to fetch F&O symbols")
        return

    # Prepare DataFrame
    df = pd.DataFrame(columns=[
        "Symbol", "60 Days Average", "Current Price", 
        "50 SMA", "20 SMA", "% diff", "Comment"
    ])

    session = create_session()
    request_counter = 0

    for symbol in fno_symbols:
        try:
            request_counter += 1
            if request_counter % 10 == 0:
                time.sleep(2)  # Add delay every 10 requests
            
            stock = yf.Ticker(symbol, session=session)
            
            # Get 60-day average
            avg_price = find_60days_historical_avg(symbol, session)
            if avg_price is None:
                continue
                
            # Get current price
            current_price = fetch_current_price(stock)
            if current_price is None:
                continue
                
            # Get SMAs
            sma_50 = fetch_x_SMA_value(stock, 50)
            sma_20 = fetch_x_SMA_value(stock, 20)
            if sma_50 is None or sma_20 is None:
                continue
                
            # Calculate difference
            price_diff = cal_diff(current_price, avg_price)
            
            # Determine comment
            if current_price > sma_20 and current_price > sma_50:
                comment = "Above 20 & 50 SMA"
            elif current_price > sma_20 and current_price < sma_50:
                comment = "Above 20 SMA, Below 50 SMA"
            elif current_price < sma_20 and current_price > sma_50:
                comment = "Below 20 SMA, Above 50 SMA"
            else:
                comment = "Below 20 & 50 SMA"
            
            df.loc[len(df)] = [
                symbol, avg_price, current_price,
                sma_50, sma_20, price_diff, comment
            ]
            
            time.sleep(1)  # Basic rate limiting
            
        except Exception as e:
            print(f"Error processing {symbol}: {e}")
            continue

    # Save to Excel
    df.to_excel("fno_analysis.xlsx", index=False, engine="openpyxl")

    # Apply formatting
    wb = load_workbook("fno_analysis.xlsx")
    ws = wb.active

    green_fill = PatternFill(start_color="00FF00", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", fill_type="solid")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        comment = row[6].value
        fill = None
        
        if comment == "Above 20 & 50 SMA":
            fill = green_fill
        elif "Above" in comment:
            fill = yellow_fill
        else:
            fill = red_fill
            
        for cell in row:
            cell.fill = fill

    wb.save("fno_analysis.xlsx")
    print("Analysis completed successfully")

if __name__ == "__main__":
    if is_last_thursday():
        main()
    else:
        print("Not the last Thursday. Exiting.")