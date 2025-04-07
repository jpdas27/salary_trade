import yfinance as yf
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta
import time
import requests
from requests.adapters import HTTPAdapter
from requests.packages.urllib3.util.retry import Retry

pd.set_option('display.float_format', '{:.2f}'.format)

def is_last_thursday():
    today = datetime.today()
    # Find last day of the month
    last_day = today.replace(day=1) + timedelta(days=32)
    last_day = last_day.replace(day=1) - timedelta(days=1)

    # Find the last Thursday
    while last_day.weekday() != 3:  # 3 = Thursday
        last_day -= timedelta(days=1)

    return today.date() == last_day.date()

def create_session():
    session = requests.Session()
    retries = Retry(
        total=5,
        backoff_factor=1,
        status_forcelist=[429, 500, 502, 503, 504]
    )
    adapter = HTTPAdapter(max_retries=retries)
    session.mount("http://", adapter)
    session.mount("https://", adapter)
    return session

stocks_list =[
    "RELIANCE.NS", "TCS.NS", "HDFCBANK.NS", "INFY.NS", "ICICIBANK.NS", "HINDUNILVR.NS", "SBIN.NS", "BHARTIARTL.NS",
    "KOTAKBANK.NS", "BAJFINANCE.NS", "ASIANPAINT.NS", "MARUTI.NS", "LT.NS", "AXISBANK.NS", "HCLTECH.NS", "WIPRO.NS",
    "M&M.NS", "NESTLEIND.NS", "ULTRACEMCO.NS", "SUNPHARMA.NS", "ITC.NS", "TECHM.NS", "TITAN.NS", "POWERGRID.NS",
    "NTPC.NS", "TATASTEEL.NS", "ONGC.NS", "BAJAJ-AUTO.NS", "INDUSINDBK.NS", "HDFCLIFE.NS", "DRREDDY.NS", "CIPLA.NS",
    "BRITANNIA.NS", "GRASIM.NS", "JSWSTEEL.NS", "ADANIPORTS.NS", "SBILIFE.NS", "EICHERMOT.NS", "HINDALCO.NS",
    "DIVISLAB.NS", "BPCL.NS", "COALINDIA.NS", "SHREECEM.NS", "IOC.NS", "TATAMOTORS.NS", "PIDILITIND.NS",
    "GODREJCP.NS", "AMBUJACEM.NS", "DABUR.NS", "HAVELLS.NS", "HEROMOTOCO.NS", "ICICIPRULI.NS", "ICICIGI.NS",
    "ADANIGREEN.NS", "ADANITRANS.NS", "ADANIENT.NS", "BANKBARODA.NS", "PNB.NS", "CANBK.NS", "UNIONBANK.NS",
    "IDBI.NS", "BANKINDIA.NS", "FEDERALBNK.NS", "IDFCFIRSTB.NS", "YESBANK.NS", "GAIL.NS", "SIEMENS.NS", "BOSCHLTD.NS",
    "MOTHERSUMI.NS", "LUPIN.NS", "BIOCON.NS", "TORNTPHARM.NS", "PAGEIND.NS", "COLPAL.NS", "BATAINDIA.NS", "VOLTAS.NS",
    "BERGEPAINT.NS", "MUTHOOTFIN.NS", "CHOLAFIN.NS", "SRF.NS", "DLF.NS", "BALKRISIND.NS", "APOLLOHOSP.NS"
]


def find_60days_historical_avg(symbol, session):
    try:
        stock_data = yf.download(symbol, period='90d', session=session)
        
        if len(stock_data) < 60:
            print(f"{symbol}: Insufficient data for 60-day calculation")
            return None
            
        filter_stock_data = stock_data.tail(60).copy()
        filter_stock_data['average_price'] = (
            filter_stock_data['Open'] + 
            filter_stock_data['High'] + 
            filter_stock_data['Low'] + 
            filter_stock_data['Close']
        ) / 4
        
        return round(filter_stock_data['average_price'].mean(), 2)
        
    except Exception as e:
        print(f"Error in 60-day avg for {symbol}: {e}")
        return None

def fetch_current_price(stock_obj):
    try:
        return round(stock_obj.fast_info["last_price"], 2)
    except Exception as e:
        print(f"Error fetching current price: {e}")
        return None

def fetch_x_SMA_value(stock_obj, days_to_cal):
    try:
        df = stock_obj.history(period="3mo")
        sma_days = f'SMA_{days_to_cal}'
        df[sma_days] = df["Close"].rolling(window=days_to_cal).mean()
        return round(df[sma_days].dropna().iloc[-1], 2)
    except Exception as e:
        print(f"Error calculating SMA {days_to_cal}: {e}")
        return None

def cal_diff(cur_price, avg_price):
    if avg_price is None or cur_price is None:
        return None
    try:
        return round(((cur_price - avg_price) / avg_price) * 100, 2)
    except ZeroDivisionError:
        return None

def main():
    # Get F&O stocks
    fno_symbols = stocks_list
    if not fno_symbols:
        print("Failed to fetch F&O symbols")
        return

    # Prepare DataFrame
    df = pd.DataFrame(columns=[
        "Symbol", "60 Days Average", "Current Price", 
        "50 SMA", "20 SMA", "% diff", "Comment"
    ])

    session = create_session()
    request_counter = 0

    for symbol in fno_symbols:
        try:
            request_counter += 1
            if request_counter % 10 == 0:
                time.sleep(2)  # Add delay every 10 requests
            
            stock = yf.Ticker(symbol, session=session)
            
            # Get 60-day average
            avg_price = find_60days_historical_avg(symbol, session)
            if avg_price is None:
                continue
                
            # Get current price
            current_price = fetch_current_price(stock)
            if current_price is None:
                continue
                
            # Get SMAs
            sma_50 = fetch_x_SMA_value(stock, 50)
            sma_20 = fetch_x_SMA_value(stock, 20)
            if sma_50 is None or sma_20 is None:
                continue
                
            # Calculate difference
            price_diff = cal_diff(current_price, avg_price)
            
            # Determine comment
            if current_price > sma_20 and current_price > sma_50:
                comment = "Above 20 & 50 SMA"
            elif current_price > sma_20 and current_price < sma_50:
                comment = "Above 20 SMA, Below 50 SMA"
            elif current_price < sma_20 and current_price > sma_50:
                comment = "Below 20 SMA, Above 50 SMA"
            else:
                comment = "Below 20 & 50 SMA"
            
            df.loc[len(df)] = [
                symbol, avg_price, current_price,
                sma_50, sma_20, price_diff, comment
            ]
            
            time.sleep(1)  # Basic rate limiting
            
        except Exception as e:
            print(f"Error processing {symbol}: {e}")
            continue

    # Save to Excel
    df.to_excel("fno_analysis.xlsx", index=False, engine="openpyxl")

    # Apply formatting
    wb = load_workbook("fno_analysis.xlsx")
    ws = wb.active

    green_fill = PatternFill(start_color="00FF00", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", fill_type="solid")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        comment = row[6].value
        fill = None
        
        if comment == "Above 20 & 50 SMA":
            fill = green_fill
        elif "Above" in comment:
            fill = yellow_fill
        else:
            fill = red_fill
            
        for cell in row:
            cell.fill = fill

    wb.save("fno_analysis.xlsx")
    print("Analysis completed successfully")

if __name__ == "__main__":
    # if is_last_thursday():
    #     main()
    # else:
    #     print("Not the last Thursday. Exiting.")
    main()

# import os
# import shutil

# output_file = "fno_analysis.xlsx"

# if os.path.exists(output_file):
#     print(f"File {output_file} exists. Moving it to GitHub workspace...")
#     shutil.move(output_file, "/github/workspace/fno_analysis.xlsx")
# else:
#     print(f"Error: File {output_file} not found. Check script execution.")